import serial
import time
import xlsxwriter
from openpyxl.workbook import Workbook
import os
import sys


ack=os.path.exists('sample.xlsx')
if(ack==0):
    headers = ['panel_id', 'zone_id', 'sensor_id', 'sensor_status', 'date', 'time']
    workbook_name = 'sample.xlsx'
    wb = Workbook()
    page = wb.active
    page.title = 'My_data'
    page.append(headers)
    wb.save(filename=workbook_name)
    wb.close()
    exit()

    #os.execv(sys.argv[0], sys.argv)





headers = ['panel_id','zone_id','sensor_id','sensor_status','date','time']
workbook_name = 'sample.xlsx'
wb = Workbook()
page = wb.active
page.title = 'PNB_data'
page.append(headers) # write the headers to the first line


#startrow = writer.sheets['Sheet1'].max_row

#SERIAL_PORT = "/dev/tty.usbserial-14255"
SERIAL_PORT = "COM6"

ser = serial.Serial(SERIAL_PORT, baudrate=9600, timeout=3)





while 1:
    serial_line = str(ser.readline())
    #serial_line = str(input())


    if serial_line != "b''" :
        md=serial_line[serial_line.find("#"):serial_line.find("@")]
        if len(md)==28 and (serial_line.find("#")+ serial_line.find("@"))==32 :
            panel_id=serial_line[3:4]
            zone_id=serial_line[5:6]
            sensor_id=serial_line[7:8]
            sensor_status=serial_line[9:10]
            date=serial_line[11:21]
            time=serial_line[22:30]
            print(panel_id,zone_id,sensor_id,sensor_status,date,time)
            print("inexcel")
            from openpyxl import load_workbook

            workbook_name = 'sample.xlsx'
            wb = load_workbook(workbook_name)
            page = wb.active

            # New data to write:
            new_companies = [[panel_id,zone_id,sensor_id,sensor_status,date,time]]

            for info in new_companies:
                page.append(info)

            wb.save(filename=workbook_name)
            wb.close()














